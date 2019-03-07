#!/usr/bin/env
#-*- coding:utf-8 -*-
import os
import json
import requests
from openpyxl import load_workbook
from openpyxl import Workbook
#地址转经纬度
def getlnglat_address(address):
    url = 'http://api.map.baidu.com/geocoder/v2/'
    output = 'json'
    ak = 'TYoptjma8Gn1fLlK4UzWcfMxAKRdbb5Y' # 百度地图ak，具体申请自行百度，提醒需要在“控制台”-“设置”-“启动服务”-“正逆地理编码”，启动
    uri = url + '?' + 'address=' + address  + '&output=' + output + '&ak=' + ak 
    response = requests.get(uri)
    temp = response.json() 
    lat = temp['result']['location']['lat']
    lng = temp['result']['location']['lng']
    #print lat,lng
    return lat,lng   # 纬度 latitude   ，   经度 longitude  ，
#经纬度转地址
def getlnglat_location(location):
    url = 'http://api.map.baidu.com/geocoder/v2/?'
    output = 'json'
    ak = 'TYoptjma8Gn1fLlK4UzWcfMxAKRdbb5Y' # 百度地图ak，具体申请自行百度，提醒需要在“控制台”-“设置”-“启动服务”-“正逆地理编码”，启动
    uri = url + 'location=' + location  + '&output=' + output + '&pois=1&ak=' + ak 
    response = requests.get(uri)
    if response.content:
        temp = response.json()
        return temp['result']['formatted_address']
    return 0
#读取excel文件中经纬度信息
def readAndWriteExcel():
    wb = load_workbook("pointList.xlsx")
    sheet = wb.get_sheet_by_name("PointAndAddress")
    modes = wb.get_sheet_by_name("totle")
    MarkersList = []
    modeList = {}
    for i in range(2,modes.max_row + 1):
        if modes["A" + str(i)].value not in modeList:
            modeList[modes["A" + str(i)].value] = int(modes["B" + str(i)].value)
        else:
            print modes["A" + str(i)].value + ":" + modeList[modes["A" + str(i)].value]

    for i in range(2,sheet.max_row + 1):
        location = sheet["A" + str(i)]
        #print location.value
        exchange = location.value.split(',')
        inLocation = exchange[1] + ',' +exchange[0]
        address = getlnglat_location(inLocation)
        if sheet["D" + str(i)].value == None:
            marker = markerName(exchange[1], exchange[0], modeList, address)
            MarkersList.append(marker[0])
            sheet["B" + str(i)].value = inLocation
            sheet["E" + str(i)].value = address
            sheet["F" + str(i)].value = marker[0]
            sheet["C" + str(i)].value = marker[1]
            sheet["D" + str(i)].value = marker[2]
            modeList = marker[3]
        #print inLocation
    j = 2
    for key in modeList.keys():
        modes["A" + str(j)] = key
        modes["B" + str(j)] = modeList[key]
        j = j + 1
    wb.save("pointList.xlsx")
    return MarkersList
#标注命名
def markerName(lat, lng, modeList, address):

    mode = checkModes(float(lat), float(lng))
    if mode not in modeList:
        modeList[mode] = 0
    modeList[mode] = modeList[mode] + 1
    markerName = mode + str(modeList[mode])
    marker = "{content:\"" + address + "\",title:\"" + markerName + "\",imageOffset: {width:-92,height:-46},position:{lat:" + lat + ",lng:" + lng + "}}"
    return marker, mode, markerName, modeList
#读取地图文件并添加标注
def readHtmlAndRevise():
    htmlf=open('map.html','rw')
    wb = load_workbook("pointList.xlsx")
    sheet = wb.get_sheet_by_name("PointAndAddress")
    newhtml = ""
    markerIn = False
    lib = {}
    
    for i in range(2,sheet.max_row + 1):
        lists = []
        lists.append(sheet["A" + str(i)].value)
        lists.append(sheet["F" + str(i)].value)
        toMarker = sheet["G" + str(i)].value
        toMarkers = []
        if toMarker == None:
            marker = "None"
            toMarkers.append(marker)
        else:
            markers = toMarker.split(",")
            for maker in markers:
                toMarkers.append(maker)
        lists.append(toMarkers)
        if sheet["D" + str(i)].value not in lib.keys():
            lib[sheet["D" + str(i)].value] = lists
        else:
            print "error: has more then one" + sheet["D" + str(i)].value

    for line in htmlf:
        newhtml = newhtml + line
        if line.strip() == "function addMapOverlay(){":
            markerIn = True
            continue
        if line.strip() == "var walking = new BMap.WalkingRoute(map);":
            for Key in lib.keys():
                pointBgein = lib[Key][0]
                pointEnds = lib[Key][2]
                if pointEnds[0] == "None":
                    continue
                for pointEndName in pointEnds:
                    #print Key + "to:" + pointEndName
                    pointEnd = lib[pointEndName][0]
                    addLine = u"\t" + "walking.search(new BMap.Point(" + pointBgein + "), new BMap.Point(" + pointEnd + "));\n"
                    #print addLine
                    newhtml = newhtml + addLine.encode("utf-8")
        if markerIn:
            #for i in range(2,sheet.max_row + 1):
            #    addLine = u"\t" + sheet["F" + str(i)].value + u",\n"
             #   newhtml = newhtml + addLine.encode("utf-8") 
            for Key in lib.keys():
                values = lib[Key][1]
                addLine = u"\t" + values + u",\n"
                newhtml = newhtml + addLine.encode("utf-8")
            markerIn = False   
    htmlf.close()
    file = r'maps.html'
    with open(file, 'w+') as f:
        f.write(newhtml) 
def checkModes(lat, lng):
    #location = "31.306952,121.159629"
    latr = lat - 31.306952
    lngr = lng - 121.159629
    if abs(latr) < 0.05 and abs(lngr) < 0.08:
        return "A"
    langlist = [int(latr*20), int(lngr*12.5)]
    print langlist
    for m in makerMode.keys():
        if makerMode[m] == langlist:
            #print m
            return m
    return "J"
#纬度偏差×10，经度偏差×10
makerMode = {
    "A":[0,0],
    "B":[1,0],
    "C":[1,-1],
    "D":[0,-1],
    "E":[-1,-1],
    "F":[-1,0],
    "G":[-1,1],
    "H":[0,1],
    "I":[1,1]
}
if __name__=='__main__':
    address = '上海市嘉定区安亭镇安拓路56号汽车·创新港'
    location = "31.306952,121.159629"
    #标签增加或删除，从新整理exlce
    resetExcel = False
    if resetExcel:
        readAndWriteExcel()
    readHtmlAndRevise()
    os.system("google-chrome maps.html")
    #checkModes(31.446952, 121.079629)